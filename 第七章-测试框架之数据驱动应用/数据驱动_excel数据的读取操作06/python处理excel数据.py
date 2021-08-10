# -*- coding:utf-8 -*-
# @Author         : 江湖一笑
# @Time           : 2021/8/10 9:18
# @Software       : Python_study
# @Python_verison : 3.7
from openpyxl import Workbook
from openpyxl.utils import get_column_letter      # 用get_column_letter得到表格列的字母编号

# # 写
wb = Workbook()
filename ='excel测试数据1.xlsx'
ws1 = wb.active
# ws1.title = 'test_sheet'
#
# for row in range(1,40):
#     ws1.append(range(600))
#
#
# ws2 = wb.create_sheet(title='Pi')
# ws2['F5'] =3.14


ws3 = wb.create_sheet(title='data')
for row in range(10, 20):
    for col in range(27, 54):
        ws3.cell(column=col, row=row, value="{0}".format(get_column_letter(col)))
print(ws3['AA10'].value)
wb.save(filename=filename)


# 读
from openpyxl import load_workbook
wb = load_workbook(filename = 'excel测试数据.xlsx')
sheet_ranges = wb['百度搜索']
print(sheet_ranges['B3'].value)